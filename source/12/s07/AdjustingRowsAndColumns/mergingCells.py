import openpyxl

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.merge_cells("A1:D3")
sheet["A1"] = "Twelve cells merged together."
sheet.merge_cells("C5:D5")
sheet["C5"] = "Two merged cells."
wb.save("merged.xslx")
