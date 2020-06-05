import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)
print(wb.create_sheet(index=0, title="First Sheet"))
print(wb.sheetnames)
print(wb.create_sheet(index=2, title="Middle Sheet"))
print(wb.sheetnames)
wb.remove_sheet(wb.get_sheet_by_name("Middle Sheet"))
wb.remove_sheet(wb.get_sheet_by_name("Sheet1"))
print(wb.sheetnames)
