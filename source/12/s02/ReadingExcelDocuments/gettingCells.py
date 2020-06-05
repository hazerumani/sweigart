import openpyxl

wb = openpyxl.load.workbook("example.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
print(sheet["A1"])
print(sheet["A1"].value)

c = sheet["B1"]
print(c.value)

print("Row " + str(c.row) + ", Column " + c.column + " is " + c.value)
print("Cell " + c.coordinate + " is " + c.value)
print(sheet["C1"].value)

print("*" * 10)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
