# 1. Import the `openpyxl` module.
import openpyxl

# 2. Call the `openpyxl.load_workbook()` function.
wb = openpyxl.load_workbook("example.xlsx")

# 3. Get a `Workbook` object.
print(wb)

# 4. Call the `get_active_sheet()` or `get_sheet_by_name()` workbook method.
print(wb.active)

# 5. Get a `Worksheet` object.
sheet = wb.active

# 6. Use indexing or the `cell()` sheet method with `row` and `column` keyword arguments.
unit = sheet.cell(1, 1)

# 7. Get a `Cell` object.
print(unit)

# 8. Read the `Cell` object's value attribute.
print(unit.value)
